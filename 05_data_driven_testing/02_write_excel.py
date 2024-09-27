import openpyxl
from openpyxl import Workbook

## writing data to excel file 
workbook=Workbook()
workbook.create_sheet("Data")

workbook.save("testdata.xlsx")

filepath="./testdata.xlsx"

workbook=openpyxl.load_workbook(filepath)

## saving data 
for r in range(1,6):
    for c in range(1,5):
        workbook['Data'].cell(r,c).value=f"value[{r}{c}]"
workbook.save("testdata.xlsx")
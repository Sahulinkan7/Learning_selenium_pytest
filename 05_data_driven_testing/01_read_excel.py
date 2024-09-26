import openpyxl

filepath="./data.xlsx"

# File...workbook...sheets...rows...cells

workbook=openpyxl.load_workbook(filepath)
sheet=workbook['Sheet1']
print(f"number of rows {sheet.max_row}")
print(f"number of columns {sheet.max_column}")

for r in range(1,sheet.max_row+1):
    for c in range(1,sheet.max_column+1):
        print(sheet.cell(r,c).value,end="  |  ")
    print()